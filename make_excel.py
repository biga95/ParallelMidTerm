import sys
import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt

ORDER = [
    ("big_words", "Bigrammi Parole"),
    ("tri_words", "Trigrammi Parole"),
    ("big_chars", "Bigrammi Caratteri"),
    ("tri_chars", "Trigrammi Caratteri"),
]

def load_times(path):
    d = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("="):
                continue
            parts = line.split()
            if len(parts) >= 2:
                key = parts[0]
                try:
                    val = float(parts[1])
                    d[key] = val
                except:
                    pass
    return d

def make_xlsx(seq, par, out_xlsx):
    wb = Workbook()
    ws = wb.active
    ws.title = "Foglio1"

    # intestazione come il tuo file
    ws["B1"] = "Sequenziale"
    ws["C1"] = "Parallelo"
    ws["D1"] = "Speed Up = Seq/Par"

    for i, (key, label) in enumerate(ORDER, start=2):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = float(seq.get(key, 0.0))
        ws[f"C{i}"] = float(par.get(key, 0.0))
        ws[f"D{i}"] = f"=B{i}/C{i}"

    # grafico barre (Sequenziale vs Parallelo)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Tempi (s)"
    chart.y_axis.title = "Secondi"

    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1+len(ORDER))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1+len(ORDER))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "F2")

    wb.save(out_xlsx)

def make_png(seq, par, out_png):
    labels = [lab for _, lab in ORDER]
    seq_vals = [seq.get(k, 0.0) for k, _ in ORDER]
    par_vals = [par.get(k, 0.0) for k, _ in ORDER]
    speed = [(s/p if p > 0 else 0.0) for s, p in zip(seq_vals, par_vals)]

    # grafico speedup (più utile nella relazione)
    plt.figure()
    plt.bar(labels, speed)
    plt.ylabel("Speedup (Seq/Par)")
    plt.xticks(rotation=20, ha="right")
    plt.tight_layout()
    plt.savefig(out_png, dpi=200)
    plt.close()

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Uso: python make_excel.py seq_times.txt par_times.txt out.xlsx out.png")
        sys.exit(1)

    seq_path = sys.argv[1]
    par_path = sys.argv[2]
    out_xlsx = sys.argv[3]
    out_png = sys.argv[4]

    seq = load_times(seq_path)
    par = load_times(par_path)

    make_xlsx(seq, par, out_xlsx)
    make_png(seq, par, out_png)
    print("OK")
